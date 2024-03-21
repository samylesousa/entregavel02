import openpyxl
wb = openpyxl.load_workbook('tabela1.xlsx')
ws = wb.active

# values = [ws.cell(row=1,column=i).value for i in range(1,ws.max_column+1)]
# print(values)

periodos = {}
tempo_execucao = {}
prioridades = {}

#fazendo um dicionario com os valores de tempo de execução e período 
for i in range(2, ws.max_row+1):
    periodos[ws.cell(row=i,column=1).value] = ws.cell(row=i,column=2).value
    tempo_execucao[ws.cell(row=i,column=1).value] = ws.cell(row=i,column=3).value
    prioridades[ws.cell(row=i,column=1).value] = ws.cell(row=i,column=4).value


# periodos=[ws.cell(row=i,column=2).value for i in range(2, ws.max_row+1)]
# tempo_execucao=[ws.cell(row=i,column=3).value for i in range(2, ws.max_row+1)]
# prioridades = [ws.cell(row=i,column=4).value for i in range(2, ws.max_row+1)]

tarefasPrioridade = sorted(prioridades)
resultado = 0

#calculando a escabilidade
for i in range(ws.max_row-1):
    resultado = resultado + (tempo_execucao[tarefasPrioridade[i]]/periodos[tarefasPrioridade[i]])

quant = ws.max_row-1

if(resultado <= (quant*((2**(1/quant))-1))):
    print("Escalonável")
else:
    print("Não escalonável")